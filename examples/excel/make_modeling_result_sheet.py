import sys
from copy import copy
import warnings
import codecs

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.datasets import load_boston
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression
from sklearn.tree import DecisionTreeRegressor
from sklearn.neighbors import KNeighborsRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error, median_absolute_error
import openpyxl as px

from nclick.excel import writing_tools as wt
from nclick.excel import cell_styles as cs


def usascii(name):
    if name.lower() == 'us-ascii':
        return codecs.lookup('utf-8')

def evaluate(y_true, y_pred):
    mean_ae = mean_absolute_error(y_true, y_pred)
    mean_se = mean_squared_error(y_true, y_pred)
    median_ae = median_absolute_error(y_true, y_pred)
    eval_result = pd.DataFrame(
            [[mean_se, mean_ae, median_ae]],
            columns=['平均二乗誤差', '平均絶対誤差', '中央絶対誤差']
            )
    return eval_result

def yyplot(y_true, y_pred, output_path):
    fig, ax = plt.subplots(1, 1, figsize=(4, 4))
    ax.scatter(y_true, y_pred)
    ax.set_xlabel('observed', fontsize=14)
    ax.set_ylabel('predict', fontsize=14)
    plt.savefig(output_path)
    plt.close()


if __name__ == '__main__':
    warnings.filterwarnings(action='ignore', module='sklearn', message='^internal gelsd')
    codecs.register(usascii)
    wb = px.Workbook()

    # dataset Sheet
    boston = load_boston()
    feats = boston.feature_names
    target = 'PRICE'
    Xy = pd.DataFrame(boston.data, columns=feats)
    Xy[target] = boston.target
    n_sample, n_feats = Xy.shape

    ws = wb.create_sheet(title='dataset')
    wt.dataframe_to_sheet(Xy, ws, start_row=1, start_col=1, index_cols_num=0)
    wt.autoresize_columns_width(ws, ratio=1.7)

    # feature correlation Sheet
    ws = wb.create_sheet(title='feature correlation')
    corr = (Xy.corr()
              .reset_index()
              .rename(columns={'index':'feature_name'}))
    wt.dataframe_to_sheet(corr,
                          ws,
                          start_row=1,
                          start_col=1,
                          index_cols_num=1,
                          number_format='0.###')

    wt.set_conditinal_format(ws,
                             cs.colorscale_02,
                             2,
                             2+n_feats,
                             2,
                             2+n_feats)
    wt.freeze_header_and_index(ws, 2, 2)
    wt.autoresize_columns_width(ws, ratio=2.5)

    # データ準備
    Xy_train, Xy_test = train_test_split(Xy, test_size=0.3, random_state=0)
    X_train, y_train = Xy_train[feats], Xy_train[target]
    X_test, y_test = Xy_test[feats], Xy_test[target]
    X_scaler = StandardScaler()
    X_scaler.fit(X_train)
    X_train = X_scaler.transform(X_train)
    X_test = X_scaler.transform(X_test)

    # モデル定義
    lr = LinearRegression()
    dt = DecisionTreeRegressor(random_state=0)
    knn = KNeighborsRegressor()
    ws = wb.create_sheet(title='evaluation')
    model_set = [[lr, 'lr', '線形回帰'],
                 [dt, 'dt', '決定木'],
                 [knn, 'knn', 'K近傍']]
    for i, (model, model_symbol, model_name) in enumerate(model_set):
        # 学習
        model.fit(X_train, y_train)

        # 予測（訓練データ）
        y_train_pred = model.predict(X_train)
        eval_result = evaluate(y_train, y_train_pred)
        filepath = f'result/{model_symbol}_train.png'
        yyplot(y_train, y_train_pred, filepath)
        wt.dataframe_to_sheet(eval_result,
                              ws,
                              start_row=14,
                              start_col=4+i*22,
                              index_cols_num=0,
                              number_format='0.000')
        wt.paste_image(ws,
                       filepath,
                       start_row=4,
                       start_col=8+i*22)


        # 予測（試験データ）
        y_test_pred = model.predict(X_test)
        eval_result = evaluate(y_test, y_test_pred)
        filepath = f'result/{model_symbol}_test.png'
        yyplot(y_test, y_test_pred, filepath)
        wt.dataframe_to_sheet(eval_result,
                              ws,
                              start_row=40,
                              start_col=4+i*22,
                              index_cols_num=0,
                              number_format='0.000')
        wt.paste_image(ws,
                       filepath,
                       start_row=30,
                       start_col=8+i*22)

        wt.merge_cells(ws, 2, 2, 4+i*22, 24+i*22)
        wt.set_value(ws, model_name, 2, 4+i*22)
        wt.set_style(ws, cs.style_03, 2, 4+i*22)
wt.merge_cells(ws, 4, 25, 2, 2)
    wt.set_value(ws, '訓練', 4, 2)
    wt.set_style(ws, cs.style_04, 4, 2)

    wt.merge_cells(ws, 30, 51, 2, 2)
    wt.set_value(ws, '試験', 30, 2)
    wt.set_style(ws, cs.style_04, 30, 2)

    wt.rotate_text(ws, 90, 30, 2)
    wt.rotate_text(ws, 90, 4, 2)

    wt.autoresize_columns_width(ws, ratio=1.5)
    wt.set_column_width(ws, 2, 5)
    wt.set_row_height(ws, 2, 25)

    # 保存
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)
    wb.save('./result/modeling_result_sheet.xlsx')